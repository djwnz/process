#!/usr/bin/env python
################################################################################
#(C) Copyright Pumpkin, Inc. All Rights Reserved.
#
#This file may be distributed under the terms of the License
#Agreement provided with this software.
#
#THIS FILE IS PROVIDED AS IS WITH NO WARRANTY OF ANY KIND,
#INCLUDING THE WARRANTY OF DESIGN, MERCHANTABILITY AND
#FITNESS FOR A PARTICULAR PURPOSE.
################################################################################
"""
@package BM2_flash
Module to create the step for writing and reading the data flash on the gas gauge
"""

__author__ = 'David Wright (david@asteriaec.com)'
__version__ = '0.1.0' #Versioning: http://www.python.org/dev/peps/pep-0386/

#
# -------
# Imports

import sys
sys.path.insert(1, '../src/')

import Tkinter as TK
import tkFileDialog as TKFD 
from functools import partial
import process_SCPI
import process_GUI
import time
import os
import openpyxl
import csv
import struct


#
# -------
# Constants
DATA_FLASH_IDS = [0,1,2,4,
                  16,17,18,19,
                  20,21,
                  32,33,34,36,37,38,39,
                  48,49,
                  56,57,58,59,
                  60,64,65,66,67,68,
                  80,81,82,83,84,85,86,88,89,
                  90,91,92,93,94,95,96,97,
                  104,105,106,107,
                  112]

#
# -------
# Classes

class Flash_Subclass:
    """
    Class contianing all of the information needed to define a subclass in the data 
    flash.
    
    @attribute subclassID  (int)  Subclass ID of this flash item.
    @attribute length      (int)  Length of the data contained within this page.
    @attribute data        (list) Bytes that contain the data flash information.
    """ 
    def __init__(self, subclass_ID, subclass_data = []):
        """
        Initialise the class.
        
        @param  subclass_ID   (int) the subclass ID of the page
        @param  page_number   (int) the page within the subclass
        """
        self.subclassID = subclass_ID
        
        if subclass_data == []:
            self.length = 0
            self.data = []
            
        else:
            self.length = len(subclass_data)
            self.data = subclass_data
        # end if
    # end def
    
    def copy(self):
        """ 
        Return a copy of the flash page
        
        @return   (Flash_Page)   a copy of this page.
        """
        
        # create a new Flash_page entity
        return_page = Flash_Subclass(self.subclassID, list(self.data))
        
        # if the data was written to it correctly
        if (return_page.length == self.length):
            # return the page
            return return_page
        
        else:
            # there was an issue
            print 'Missmatched data lengths'
            return None
        # end if
    # end def
    
    def insert(self, offset, data):
        """ 
        Insert data into the subclass at a given offset
        
        @param    offset   (int)    The offset in the subclass to write to
        @param    data     (list)   The data to write
        """
        for item in data:
            try:
                self.data[offset] = item
            
            except:
                print 'Subclass ' + str(self.subclassID) + ', offset ' + str(offset) + ' is an overrun'
            # end try
            offset += 1
        # end for
    # end def
    
    def to_pages(self):
        """
        Convert the subclass to a list of pages to be written to data flash
        
        @return   (list)    list of Flash_Pages that make up the subclass
        """
        
        # initialise the return list
        return_pages = []
               
        # dummy variable to manage the length
        remaining_length = self.length
        current_page = 1;
        
        # store all the pages
        while remaining_length > 32:
            # initialise the flash page
            new_page = Flash_Page(self.subclassID, current_page)
            
            # define the indicies for the data
            start_index = (current_page-1)*32
            end_index = start_index + 32
            
            # add the data to the page
            new_page.append(self.data[start_index:end_index])
            
            # store the page
            return_pages.append(new_page)
            
            # 
            current_page += 1
            remaining_length -= 32
        # end while
        
        # initialise the flash page
        new_page = Flash_Page(self.subclassID, current_page)
        
        # define the indicies for the data
        start_index = (current_page-1)*32
        
        # add the data to the page
        new_page.append(self.data[start_index:])
        
        # store the page
        return_pages.append(new_page)       
        
        return return_pages
    # end def
    
    def is_equal(self, other_page):
        """
        determine if this page is equal to the other page provided.
        
        @param   other_page (Flash_Page)   The other page to compare to.
        @return  (bool)                    True if the classes are equal.
        """
        
        # compare subclass IDs
        if self.subclassID != other_page.subclassID:
            return False
        
        # compare lengths
        elif self.length != other_page.length:
            return False
        
        # compare the length of data lists
        elif len(self.data) != len(other_page.data):
            return False
        # end if
        
        # iterate through data elements comparing values
        for i in range(len(self.data)):
            if self.data[i] != other_page.data[i]:
                return False
            # end if
        # end for
        
        # If it makes it here then the pages are equal.
        return True
    # end def
    
    def append(self, additional_data):
        """
        Append new data to the page
        
        @param    additional_data  (list)  The list of bytes to append
        @return   (list)                   Any overflowing data, 
                                           None if there is none.
        """
    
        # determine the number of bytes the new data contains
        additional_length = len(additional_data)
        
        self.length += additional_length
        self.data.extend(additional_data)
    # end def
    
    def to_list(self):
        """
        Function to convert the page to a list for veiwing.
        """
        return [self.subclassID, self.length] + [self.data]
    # end def
# end class

class Flash_Page:
    """
    Class contianing all of the information needed to define a page in the data 
    flash.
    
    @attribute subclassID  (int)  Subclass ID of this flash page.
    @attribute page_number (int)  Page number within this subclass of data flash.
    @attribute length      (int)  Length of the data contained within this page.
    @attribute data        (list) Bytes that contain the data flash information.
    """ 
    def __init__(self, subclass_ID, page_number, page_data = []):
        """
        Initialise the class.
        
        @param  subclass_ID   (int) the subclass ID of the page
        @param  page_number   (int) the page within the subclass
        """
        self.subclassID = subclass_ID
        self.page_number = page_number
        
        if page_data == []:
            self.length = 0
            self.data = []
            
        else:
            self.length = len(page_data)
            self.data = page_data
        # end if
    # end def
    
    def copy(self):
        """ 
        Return a copy of the flash page
        
        @return   (Flash_Page)   a copy of this page.
        """
        
        # create a new Flash_page entity
        return_page = Flash_Page(self.subclassID, self.page_number, self.data)
        
        # if the data was written to it correctly
        if (return_page.length == self.length):
            # return the page
            return return_page
        
        else:
            # there was an issue
            print 'Missmatched data lengths'
            return None
        # end if
    # end def
    
    def is_equal(self, other_page):
        """
        determine if this page is equal to the other page provided.
        
        @param   other_page (Flash_Page)   The other page to compare to.
        @return  (bool)                    True if the classes are equal.
        """
        
        # compare subclass IDs
        if self.subclassID != other_page.subclassID:
            return False
        
        # compare lengths
        elif self.length != other_page.length:
            return False
        
        # compare page numbers
        elif self.page_number != other_page.page_number:
            return False
        
        # compare the length of data lists
        elif len(self.data) != len(other_page.data):
            return False
        # end if
        
        # iterate through data elements comparing values
        for i in range(len(self.data)):
            if self.data[i] != other_page.data[i]:
                return False
            # end if
        # end for
        
        # If it makes it here then the pages are equal.
        return True
    # end def
    
    def append(self, additional_data):
        """
        Append new data to the page
        
        @param    additional_data  (list)  The list of bytes to append
        @return   (list)                   Any overflowing data, 
                                           None if there is none.
        """
        # determine the number of bytes that can fit in this page
        fitting_length = 32 - self.length
    
        # determine the number of bytes the new data contains
        additional_length = len(additional_data)
        
        # Determine if the new data will fit in this page.
        if additional_length <= fitting_length:
            # yes it will fit so append the data and updte the length
            self.length += additional_length
            self.data.extend(additional_data)
            
            # there is no overflow
            return None
        
        else:
            # there is overflow, so add the data that does fit.
            self.data.extend(additional_data[:fitting_length])
            self.length += fitting_length
            
            # return the overflowing data to be written to the next page
            return additional_data[fitting_length:]
        # end if
    # end def
    
    def to_list(self):
        """
        Function to convert the page to a list for veiwing.
        """
        return [self.subclassID, self.page_number, self.length] + [self.data]
    # end def
    
    def to_SCPI(self):
        """ 
        Function to convert the page to a usable SCPI command
        """
        return "BM2:BQF WRITE," + str(self.subclassID) + ',' + str((self.page_number-1)*32) + ',' + str(self.length) + ',' + ','.join(self.data)
    # end def
# end class
        

class Update_Flash:
    """
    Template for a class that describes a step in a process.
    Elelments included in this template must be preserved.
    
    Essential attributes: 
    @attribute properties    (SUP_process.process_properties) 
                                             Denotes propeties of the step
    @attribute title         (String)        Title text to display
    @attribute text          (String)        Body text to display
    @attribute Header        (TK label)      The title as displayed in the GUI
    @attribute body          (TK label)      Body text display in GUI
    @attribute error_label   (TK label)      Text to display if there is an error
    @attribute no_aardvark   (bool)          True is a connection to an aardvark 
                                             has been made, False otherwise.
                                             
    Other attributes can be added as needed
    """
    def __init__(self, properties):
        """
        Initialise the Class
        
        @param  properties   (SUP_process.process_properties) 
                                             Denotes propeties of the step
        """
        # define the title of the step
        self.title = "Update Flash"
        
        # define the body text for the step
        self.text = "Updates the data flash or the battery module gas gauge"
        
        # initialise attributes 
        self.no_aardvark = False
        # include other attributes as needed
        
        # copy the properties
        self.properties = properties
        
        # initialise path to the excel document to parse
        self.excel_path = ''
        
        # initialise the name of the configuration to use
        self.configuration = ''
        
        # initialise the dict of parsed flash pages
        self.parsed_flash_pages = {}
        
        # initialise the dict of read flash pages
        self.read_flash_pages = {}    
        
    # end def
    
    def disable_buttons(self):
        """
        Disable all buttons.
        """
        self.load_button.config(state = 'disabled')
        self.write_button.config(state = 'disabled')
        self.read_button.config(state = 'disabled')  
        self.parse_button.config(state = 'disabled')
        self.save_button.config(state = 'disabled')
    # end def
    
    def load_csv(self):
        """
        Load saved data flash dictionary from a saved csv file
        """  
        
        # prepare to open a window to load a file through
        file_opt = options = {}
        options['defaultextension'] = '.csv'
        options['filetypes'] = [('csv files', '.csv')]
        options['initialdir'] = os.getcwd()
        options['title'] = 'Select .csv file to open' 
              
        # open window to fet filename to open
        filename = TKFD.askopenfilename(**file_opt)   
        
        # see if the user selected a file or not
        if (filename != ''):  
            try:
                # open the csv file
                csv_input = open(filename, 'rb')
                
                file_text = csv_input.read()
                
                if ("# Published through BM2_flash.py" in file_text):
                    # file is compatible with this program
                    
                    # re-open the csv file
                    csv_input.close()
                    csv_input = open(filename, 'rb')                    

                    # define csv reader object
                    reader = csv.reader(csv_input, delimiter = '\t')
                    
                    # empty the read dictionary
                    self.read_flash_pages = {}
                    
                    for row in reader:
                        if (row[0] == 'Subclass ID') or row[0].startswith('#'):
                            # this is the header row
                            next
                            
                        else:
                            # this is a data row so store the data
                            self.read_flash_pages[str(row[0])] = Flash_Subclass(row[0], row[2:])
                        # end if
                    # end for
                    
                else:
                    print "This csv file is not compatible with this program"
                # end if
                
                csv_input.close()
                
                self.parse_button.config(state = 'normal')
                self.save_button.config(state = 'normal')
            
            except:
                print "Could not open and read to the csv file"
            # end try
        
        else:
            print "No file was selected to write to"
        # end if
    # end def            
    
    def save_csv(self):
        """
        Saves the loaded data flash to a csv file. If as excel file has not been
        parsed then it will save the read data flash. Otherwise it will save the
        Parsed data flash
        """      
        
        # open window for saving the file
        file_opt = options = {}
        options['defaultextension'] = '.csv'
        options['filetypes'] = [('csv files', '.csv')]
        options['initialdir'] = os.getcwd()
        options['initialfile'] = 'test.csv'
        options['title'] = 'Save .csv file as:'     
        
        # get the file name from the user
        filename = TKFD.asksaveasfilename(**file_opt)
        
        # see if the user selected a file or not
        if (filename != ''):    
            # a file was selected so open file for writing   
            try:
                # open the csv file
                csv_output = open(filename, 'wb')
                
                # write a comment line
                csv_output.write("# Published through BM2_flash.py\n")
                
                # define a csv writer object
                output_writer = csv.writer(csv_output, delimiter = '\t')
                
                # write the title
                output_writer.writerow(["Subclass ID", "Length", "Data Array ->"])
                
                if self.parsed_flash_pages.keys() == []:
                    # the parsed dictionary is empty
                    key_list = sorted_key_list(self.read_flash_pages)
                    
                    for key in key_list:
                        output_list = self.read_flash_pages[key].to_list()[:2]
                        output_list.extend(self.read_flash_pages[key].to_list()[2])
                        output_writer.writerow(output_list)
                    # end for
                    
                else:
                    key_list = sorted_key_list(self.parsed_flash_pages)
                    
                    for key in key_list:
                        output_list = self.parsed_flash_pages[key].to_list()[:2]
                        output_list.extend(self.parsed_flash_pages[key].to_list()[2])
                        output_writer.writerow(output_list)                        
                    # end for          
                # end if
                
                csv_output.close() 
                
            except:
                print "Could not open and write to the csv file"
            # end try
            
        else:
            print "No file was selected to write to"
        # end if
    # end def
    
    def write_data_flash(self):
        """
        Write the data flash subclasses to the gas gauge
        """
        
        self.disable_buttons()
        
        # initialise the list of pages to write
        pages_to_write = []
        
        # sort the keys
        keylist = sorted_key_list(self.parsed_flash_pages)

        # populate the list of pages
        for key in keylist:
            if not self.parsed_flash_pages[key].is_equal(self.read_flash_pages[key]):
                pages_to_write.extend(self.parsed_flash_pages[key].to_pages())
        # end for
        
        with process_SCPI.aardvark() as AARD:
            # initialise the aardvark
            if AARD.port != None:
                
                # get the serial number
                serial_number = AARD.read_SCPI("SUP:TEL? 9,data", 
                                             self.properties.address, 
                                             'uint')
                
                # unlock the flash
                AARD.send_SCPI(("SUP:NVM UNLOCK," + str(serial_number+12345)), self.properties.address) 
                
                # write the flash updating commands
                for page in pages_to_write:
                    AARD.send_SCPI(page.to_SCPI(), self.properties.address) 
                    time.sleep(1)
                # end for
                
                # lock the flash
                AARD.send_SCPI("SUP:NVM WRITE,1", self.properties.address)                            
                
            else:
                # no aardvark was found
                self.no_aardvark = True
                
            #end if
        # end with        
        
        # re-enable buttons
        self.parse_button.config(state = 'normal')
        self.save_button.config(state = 'normal')
        self.load_button.config(state = 'normal')
        self.write_button.config(state = 'normal')
        self.read_button.config(state = 'normal')      
    # end def
            
    
    def read_data_flash(self):
        """
        Read the data flash from the gas gauge
        """
        
        self.disable_buttons()
        
        for ID in DATA_FLASH_IDS:
            current_page = 1
            
            recieved_subclass = Flash_Subclass(ID)
            
            recieved_page = self.read_BM2_page(ID, current_page)
            
            if (recieved_page.length == 0):
                recieved_page = self.read_BM2_page(ID, current_page)
                if (recieved_page.length == 0):
                    print "subclass " + str(ID) + " is empty"
                    
                else:
                    # store the page
                    recieved_subclass.append(recieved_page.data)
                # end if
            
            else:
                # store the page
                recieved_subclass.append(recieved_page.data)
            # end if
            
            while recieved_page.length == 32:
                # look for extra pages
                current_page += 1
                recieved_page = self.read_BM2_page(ID, current_page)
                
                if (recieved_page.length == 0):
                    # catch for empty pages
                    pass
                
                else:
                    # store the page
                    recieved_subclass.append(recieved_page.data)
                # end if
            # end while
            
            if (recieved_subclass.length != 0):
                self.read_flash_pages[str(ID)] = recieved_subclass
            # end if
        # end for
        
        self.load_button.config(state = 'normal')
        self.read_button.config(state = 'normal')  
        self.parse_button.config(state = 'normal')
        self.save_button.config(state = 'normal')
    # end def
    
    def read_BM2_page(self, ID, page):
        """
        Read a page of data from the BM2
        
        @param    ID      (int)         The subclass ID to read
        @param    page    (int)         The page within that subclass to read
        @return   (Flash_Page)          The recieved data
        """
        
        # initialise the page
        read_page = Flash_Page(ID, page)
        
        # define the SPCI command to send
        SCPI_command = "BM2:BQF READ_PAGE," + str(read_page.subclassID) + ',' + str(read_page.page_number)
        
        with process_SCPI.aardvark() as AARD:
            # initialise the aardvark
            if AARD.port != None:
                # if an aardvark was found, request the page to be read
                AARD.send_SCPI(SCPI_command, self.properties.address)
                
                # wait for 0.1 seconds
                time.sleep(0.1)
                
                # read the page
                page_data = AARD.read_SCPI("BM2:TEL? 78,data", 
                                             self.properties.address, 
                                             33*['char'])
                
                # extract the data from the result
                
                if page_data != None:
                    # the read was successful
                    read_page.length = page_data[0]
                    read_page.data = [str(item) for item in page_data[1:read_page.length+1]]
                # end if                   
                
            else:
                # no aardvark was found
                self.no_aardvark = True
                
            #end if
        # end with 
        
        return read_page
    # end def    
        
    
    def read_flash_excel(self):
        """
        Function to open an excel spreadsheet and parse it into a list of 
        pages to write to the gas gauge data flash
        """
        
        self.disable_buttons()
        
        # copy the data flash list
        for key in self.read_flash_pages:
            self.parsed_flash_pages[key] = self.read_flash_pages[key].copy()
        # end for      
        
        # get the filename of the excel spreadsheet to parse
        self.get_filename()
        if self.excel_path == '':
            print 'No Excel file selected'
            self.load_button.config(state = 'normal')
            self.read_button.config(state = 'normal')
            self.parse_button.config(state = 'normal')    
            self.save_button.config(state = 'normal')  
            return 0
        # end if
            
        # open the file
        workbook = openpyxl.load_workbook(filename = self.excel_path, data_only = True)
        
        # get the desired configuration information from the file
        self.get_flash_configuration(workbook)
        
        if (self.configuration == ''):
            # no configuration information was found
            print 'No configuration selected'
            self.load_button.config(state = 'normal')
            self.read_button.config(state = 'normal')
            self.parse_button.config(state = 'normal')            
            return 0
        # end if
        
        # extract the data flash information in entirety
        self.extract_flash_data(workbook)     
        
        # re-enable the buttons
        self.load_button.config(state = 'normal')
        self.write_button.config(state = 'normal')
        self.read_button.config(state = 'normal')
        self.parse_button.config(state = 'normal')
        self.save_button.config(state = 'normal')
    # end def
    
    def get_flash_configuration(self, workbook):
        """
        Extract the possible configurations for the workbook and then ask the
        user to select one of them
        
        @param    workbook  (openpyxl.workbook) The workbook to examine
        """
        # Open the first cell in the first page
        first_sheet = workbook[workbook.sheetnames[0]]
        first_cell = first_sheet['A1']
        
        # check to see if you have a non-data page
        if first_cell.value != 'Subclass ID':
            # this is a non-data page so pick the next page.
            first_sheet = workbook[workbook.sheetnames[1]]   
            first_cell = first_sheet['A1']
        # end if
        
        # initialise the list of headers
        config_titles = []
        
        # iterate through the first row
        for cell in first_sheet['1']:
            # check to see if this is a configuration title
            if 'Cell' in str(cell.value):
                # it is so append to the list
                config_titles.append(cell.value)
            # end if
        # end for
        
        # define a pop-up window
        self.config_window = TK.Toplevel()
        self.config_window.title('Select the desired Configuration')
        
        # define the options variable for the Menu
        self.selected_option = TK.StringVar()
        self.selected_option.set(config_titles[0])
        
        # define the Option Menu to allow configuration selection
        config_menu = apply(TK.OptionMenu, (self.config_window, self.selected_option) + tuple(config_titles))
        config_menu.grid(row = 0, column = 0)
        
        # define the button to execute selection of the configurtion
        select_button = TK.Button(self.config_window, text = 'Select', 
                                 command = self.set_configuration, 
                                 activebackground = 'green', width = 15)
        select_button.grid(row = 0, column = 1)
        
        # clear the current configuration
        self.configuration = ''
        
        # wait for the pop-up window to be exited
        self.config_window.wait_window()    
        
        print self.configuration
    # end def
    
    def extract_flash_data(self, workbook):
        """
        Perform the extraction of the flash data from the spreadsheet
        
        @param workbook  (openpyxl.workbook) the workbook to extract from
        """
        
        # iterate through the worksheets in the workbook
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # skip non-data sheets
            first_cell = worksheet['A1']
            if first_cell.value != 'Subclass ID':
                continue
            # end if
            
            # read in the row of headers
            header_row = worksheet['1']
            
            # convert to ascii strings
            header_titles = []
            for cell in header_row:
                try:
                    header_titles.append(cell.value.encode('ascii','ignore'))
                    
                except:
                    header_titles.append('empty')
                # end try
            # end if
            
            # get the indicies of the relevant columns
            format_column = header_titles.index('Format')
            size_column = header_titles.index('Size')
            data_column = header_titles.index(self.configuration)
            type_column = header_titles.index('Type')
            offset_column = header_titles.index('Offset')
            
            # initialise variables
            row_to_parse = 2
            data_valid = True
            current_subclass_ID = -1
            current_page = 1
            parsing_page = None
            
            # while you are still successfully parsing data
            while data_valid:
                
                # extract all of the contents of the row
                row = []
                for cell in worksheet[str(row_to_parse)]:
                    value_type = type(cell.value)
                    
                    if value_type in [int, long]:
                        row.append(int(cell.value))
                        
                    elif value_type is float:
                        row.append(cell.value)
                        
                    elif value_type in [str, unicode]:
                        row.append(cell.value.encode('ascii','ignore'))
                        
                    else:
                        row.append(str(value_type))
                    # end if
                # end for
                
                # pull out the items that are wanted
                subclass_ID = row[0]
                data_format = row[format_column]
                data_size = row[size_column]
                data = row[data_column]
                data_type = row[type_column]
                offset = row[offset_column]
                
                if ((type(subclass_ID) != int) or 
                    (type(data_size) != int) or 
                    (type(offset) != int) or 
                    (type(data_format) != str)):
                    
                    # this row has in-valid data, this worksheet is finished
                    data_valid = False
                    
                    break
                # end try

                # check if this data item is protected
                if ((data_type == 'Unknown') or
                    (data_type == 'Lifetime' and not self.lifetime_erase.get()) or
                    (data_type == 'Calibration' and not self.cal_erase.get())):
                    next
                    
                else:
                    # this item is not protected
                    
                    # parse the information from this row into a new data list.
                    data_list = self.parse_row(data_size, data_format, data)  
                    
                    # overrite the previous data
                    self.parsed_flash_pages[str(subclass_ID)].insert(offset, data_list)
                # end if
                
                row_to_parse += 1
            # end while
        # end for
        
        keylist = sorted_key_list(self.read_flash_pages)
        
        for key in keylist:
            if not self.read_flash_pages[key].is_equal(self.parsed_flash_pages[key]):
                print "Subclass " + key + " Has Changes" 
                print "Originally\t= " + str(self.read_flash_pages[key].to_list())
                print "Now\t\t= " + str(self.parsed_flash_pages[key].to_list())
                
            else:
                print "Subclass " + key + " is unchanged and had length " + str(self.read_flash_pages[key].length)
            # end if
    # end def
    
    def parse_row(self, size, format_string, data):
        """
        Parse the data collected from a row into a list of bytes
        
        @param   size          (int)      The number of bytes to convert into
        @param   format_string (string)   The format of the data
        @param   data          (variable) The data to parse
        @return  (list)                   The data parsed into a list of bytes
        """
        
        try:
            if format_string in ['Integer', 'Unsigned Integer']:
                # define the formt string for the appropriate length int
                format_str = '{:0' + str(size*2) + 'X}'
                
                # convert the data number to hex of the appropriate length
                hex_string = format_str.format(data & (2**(8*size)-1))
                
                # convert each pair of hex characters to an int and build the list
                return_list = []
                for i in range(size):
                    return_list.append(str(int(hex_string[i*2:i*2+2],16)))
                # end for
                
                return return_list  
            
            elif format_string == 'Hex':
                # strip off the 0x st the beginning
                trimmed_hex = data[2:]
                return_list = []
                
                # convert each pair of hex characters to an int and build the list
                for i in range(size):
                    return_list.append(str(int(trimmed_hex[i*2:i*2+2],16)))
                # end for
                
                return return_list
            
            elif format_string == 'String':
                if data == "<type 'NoneType'>":
                    return size*['0']
                
                else:
                    # find the string length
                    string_len = len(data)
                    
                    # format the list with the string length first and then pad the end 
                    # with zeros
                    return [str(string_len)] + [str(ord(c)) for c in data] + (size-string_len-1)*['0']                  
                # end if
                
            elif format_string == 'Float':
                return_list = [str(b) for b in encode_TI_float(data)]
                return return_list
                
            else:
                print 'Unrecognised Format = ' + format_string
                return size*['0']
            # end if
        
        except:
            print 'failed to parse ' + format_string + ' ' + str(data)
            return size*['0']
    # end def
        
    def set_configuration(self):
        """
        Take provided configuration and store it to the class
        
        """
        self.configuration = self.selected_option.get()
        self.config_window.destroy()
    # end def
    
    
    def get_filename(self):
        """
        Prompt the user to select the file to poen
        """
        
        # prepare to open a window to load a file through
        file_opt = options = {}
        options['defaultextension'] = '.xlsx'
        options['filetypes'] = [('xlsx files', '.xlsx')]
        options['title'] = 'Select data flash file to open' 
        
        # determine default directory to open
        if os.path.isdir("C:\Pumpkin\BM\BQ34Z651_Files"):
            options['initialdir'] = "C:\Pumpkin\BM\BQ34Z651_Files"
            dir_list = os.listdir("C:\Pumpkin\BM\BQ34Z651_Files")
            
        else:
            dir_list = os.listdir(os.getcwd())
        # end if
        
            # use the name of the first file in the directory
            options['initialfile'] = dir_list[1]
        # end if
        
        # open window to fet filename to open
        filename = TKFD.askopenfilename(**file_opt)
         
        # store the filename
        self.excel_path = filename
    # end def        
    
    def execute(self):
        """
        The code to run for this step, must be present, may be nothing.
        
        This function runs before the GUI is loaded, while the please wait 
        message is displayed.
        
        If the step uses an Aarvark, attempt to communicate with it here and 
        set self.no_aardvark if this communication fails.
        """
        # trivial execution 
        self.no_aardvark = False
    # end def
    
    def gui(self, parent_frame):
        """
        This function lodas the gui elements associated with this step.
        
        @param parent_frame  (TK Frame)   The area in the GUI where these 
                                          elements can be placed.
        """
        # the step title
        self.Header = TK.Label(parent_frame, text = self.title)
        self.Header.config(font = process_GUI.title_font, 
                           bg = process_GUI.default_color)
        self.Header.grid(row = 0, column=0, columnspan = 2, sticky = 'nsew')
        
        # the body text for the step
        self.body = TK.Label(parent_frame, text = self.text)
        self.body.config(font = process_GUI.label_font, 
                           bg = process_GUI.default_color)
        self.body.grid(row = 1, column=0, columnspan = 2, sticky = 'nsew')
        
        # button to trigger reading of dataflash from the gas gauge
        self.read_button = TK.Button(parent_frame, text = 'Read Data Flash', 
                                          command = self.read_data_flash, 
                                          activebackground = 'green', width = 15)
        self.read_button.grid(row = 2, column = 0, sticky = 'nsew')      
        
        # button to trigger loading of dataflash from a csv file
        self.load_button = TK.Button(parent_frame, text = 'Load Data Flash', 
                                          command = self.load_csv, 
                                          activebackground = 'green', width = 15)
        self.load_button.grid(row = 2, column = 1, sticky = 'nsew')         
        
        # checkbox to determine if lifetime data should be erased
        self.lifetime_erase = TK.IntVar()
        self.lifetime_checkbox = TK.Checkbutton(parent_frame, text = 'Erase Lifetime Information',
                                                variable = self.lifetime_erase)
        self.lifetime_checkbox.grid(row = 3, column = 0, columnspan = 2, sticky = 'nsew')
        
        # checkbox to determine if calibration data should be erased
        self.cal_erase = TK.IntVar()
        self.cal_checkbox = TK.Checkbutton(parent_frame, text = 'Erase Calibration Information',
                                                variable = self.cal_erase)
        self.cal_checkbox.grid(row = 4, column = 0, columnspan = 2, sticky = 'nsew')
        
        # the button to trigger loading of information from the excel file
        self.parse_button = TK.Button(parent_frame, text = 'Load Excel File', 
                                         command = self.read_flash_excel, 
                                         activebackground = 'green', width = 15,
                                         state = 'disabled')
        self.parse_button.grid(row = 5, column = 0, columnspan = 2, sticky = 'nsew')        
        
        # button to trigger writing of dataflash to the gas gauge
        self.write_button = TK.Button(parent_frame, text = 'Write Data Flash', 
                                     command = self.write_data_flash, 
                                     activebackground = 'green', width = 15,
                                     state = 'disabled')
        self.write_button.grid(row = 6, column = 0, sticky = 'nsew')  
        
        # button to trigger saving of dataflash to a csv file
        self.save_button = TK.Button(parent_frame, text = 'Save Data Flash', 
                                     command = self.save_csv, 
                                     activebackground = 'green', width = 15,
                                     state = 'disabled')
        self.save_button.grid(row = 6, column = 1, sticky = 'nsew')         
        
        # This check is only required if the step 
        # requires the use of the arrdvark
        if self.no_aardvark:
            # if arrdvark communications were unsuccessful print an error
            self.error_label = TK.Label(parent_frame, 
                                        text = "No Aardvark Connected")
            self.error_label.config(font = process_GUI.label_font, 
                                 bg = process_GUI.default_color)
            self.error_label.grid(row = 2, column=0, columnspan = 2)   
        
        else:
            # Aardvark communications were successful so load the rest of the GUI
            # add step specific GUI elements here
            next
    #end def
    
    def close(self):
        """
        Deconstruct the GUI for this step
        """
        # remove the fixed items
        self.Header.grid_forget()
        self.body.grid_forget()
        self.load_button.grid_forget()
        self.lifetime_checkbox.grid_forget()
        self.cal_checkbox.grid_forget()
        self.write_button.grid_forget()
        self.read_button.grid_forget()
        self.save_button.grid_forget()
        self.parse_button.grid_forget()
        
        if self.no_aardvark:
            # if an aardvark was not found remove the error label
            self.error_label.grid_forget()
            
        else:
            # if an aardvark was found remove all other items
            # grid_forget all step specific GUI elements here so that they 
            # are removed from the GUI
            next
        # end if
    #end def
#end class

def sorted_key_list(input_dict):
    """ 
    Function to sort the keys of a dictionary into numerical order.
    
    @param    input_dict   (Dictionary)   The dictionary to sort.
    @return   (list)                The orted list of keys.
    """
    
    # define relevant lists
    number_list = []
    key_list = input_dict.keys()
    bad_key_list = []
    
    # terate through the keys and convert to integers
    for key in key_list:
        try:
            number_list.append(int(key))
        
        except:
            print key + ' is not a valid number to sort'
            bad_key_list.append(key)
        # end try
    # end for
    
    #sort the list
    number_list.sort()
    
    # convert back to strings
    key_list = [str(number) for number in number_list]
    
    # add the bad keys
    key_list.extend(bad_key_list)
    
    return key_list
# end def
            
def encode_TI_float(val):
    """
    Encode a floating point number as per TI's encoding format as per page 119
    of sluuax0c.pdf.
    
    @param    val         (float)   The floating point value to encode.
    @return   rawData     (list)    The encoded float as a four byte list
    """
    rawData = 4*[0]
    
    exp = 0
    
    if (val == 0.0):
        # catch the zero trap before getting stuck in infinite loop
        return rawData
    
    elif (val < 0):
        mod_val = -val
        
    else:
        mod_val = val
    # end if
    
    tmpVal = mod_val
    
    tmpVal = tmpVal * (1 + 2**-35)
    
    if (tmpVal < 0.5):
        while (tmpVal < 0.5):
            tmpVal*=2
            exp-=1
        # end while
        
    elif (tmpVal >= 1.0):
        while (tmpVal >= 1.0):
            tmpVal/=2.0
            exp+=1
        # end while
    # end if
    
    if (exp > 127):
        exp = 127
        
    elif (exp < -128):
        exp = -128
    # end if
    
    tmpVal = ((2**(8-exp)) * mod_val) - 128
    
    byte2 = int(tmpVal // 1)
    
    tmpVal = (2**8) * (tmpVal - byte2)
    
    byte1 = int(tmpVal // 1)
    
    tmpVal = (2**8) * (tmpVal - byte1)
    
    byte0 = int(tmpVal // 1)
    
    if (val < 0):
        byte2 |= 0x80
    # end if
    
    rawData[0] = exp + 128
    
    rawData[1] = byte2
    
    rawData[2] = byte1
    
    rawData[3] = byte0
    
    return rawData
# end def

def decode_TI_float(rawData):
    """
    Decode a floating point number as per TI's encoding format as per page 119
    of sluuax0c.pdf.
    
    @param    rawData         (list)   The list to decode.
    @return   val             (float)  The decoded float.
    """
    
    byte0 = rawData[3]
    
    byte1 = rawData[2]
    
    byte2 = rawData[1]
    
    exp = rawData[0] - 128   
    
    if (byte2 & 0x80):
        is_neg = True
        byte2 &= 0x7F
        
    else:
        is_neg = False
    # end if
    
    tmpVal = byte0
    
    tmpVal = tmpVal/(2.0**8) + byte1
    
    tmpVal = tmpVal/(2.0**8) + byte2
    
    mod_val = (tmpVal + 128)/(2.0**(8-exp))
    
    if is_neg:
        val = -mod_val
            
    else:
        val = mod_val
    # end if
    
    return val
# end def